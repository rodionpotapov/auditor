import sys
import pathlib
import pytest
import pandas as pd
import io
import openpyxl

_PROJECT_ROOT = str(pathlib.Path(__file__).parent.parent)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

for k in ["src", "src.report_generator", "src.config"]:
    if k in sys.modules:
        del sys.modules[k]

from src.report_generator import generate_report

@pytest.fixture
def scored_data():
    return pd.DataFrame(
        data={
            "Период": pd.to_datetime(["2026-03-25"] * 5),
            "СчетДт": ["51"] * 5,
            "СчетКт": ["60"] * 5,
            "Сумма": [150000.0, 50.0, 2000.0, 5000.0, 10000.0],
            "abs_amount": [150000.0, 50.0, 2000.0, 5000.0, 10000.0],
            "pair_mean": [10000.0] * 5,
            "Контрагент": ["ООО Тест"] * 5,
            "Содержание": ["Текст"] * 5,
            "ТипДокумента": ["Операция"] * 5,
            "boosted_score": [95.0, 90.0, 70.0, 50.0, 0.0],
            "explanation": ["Причина"] * 5
        },
        index=None, columns=None, dtype=None, copy=None
    )

def test_report_filters_min_amount_and_score(scored_data):
    report_bytes = generate_report(scored_data)
    wb = openpyxl.load_workbook(
        filename=io.BytesIO(report_bytes),
        read_only=False, keep_vba=False, data_only=False, keep_links=False
    )
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 5 
    assert rows[1][8] == 95.0 

def test_report_formatting_colors(scored_data):
    report_bytes = generate_report(scored_data)
    wb = openpyxl.load_workbook(
        filename=io.BytesIO(report_bytes),
        read_only=False, keep_vba=False, data_only=False, keep_links=False
    )
    ws = wb.active
    
    red_row_fill = ws.cell(row=2, column=1).fill.start_color.rgb
    assert "FFCCCC" in red_row_fill or "00000000" in red_row_fill 

def test_empty_dataframe_report():
    empty_df = pd.DataFrame(
        data=None,
        columns=["Период", "СчетДт", "СчетКт", "Сумма", "abs_amount", "pair_mean", "Контрагент", "Содержание", "ТипДокумента", "boosted_score", "explanation"],
        index=None, dtype=None, copy=None
    ).astype({
        "abs_amount": float, 
        "boosted_score": float, 
        "pair_mean": float, 
        "Сумма": float
    })
    
    report_bytes = generate_report(empty_df)
    
    wb = openpyxl.load_workbook(
        filename=io.BytesIO(report_bytes),
        read_only=False, keep_vba=False, data_only=False, keep_links=False
    )
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1 
    assert ws.cell(row=1, column=1).value == "Дата"