import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
from src.config import MIN_AMOUNT, REPORT_TOP_N


def generate_report(data: pd.DataFrame, top_n: int = REPORT_TOP_N) -> bytes:
    """Генерирует Excel отчёт в памяти и возвращает байты для скачивания."""

    # Топ аномалий с фильтром по минимальной сумме
    report_data = (
        data[data["abs_amount"] >= MIN_AMOUNT]
        .nlargest(top_n, "boosted_score")
        [[
            "Период", "СчетДт", "СчетКт", "Сумма", "pair_mean",
            "Контрагент", "Содержание", "ТипДокумента", "boosted_score", "explanation"
        ]]
        .copy()
    )

    report_data.columns = [
    "Дата", "Счет Дт", "Счет Кт", "Сумма", "Средняя сумма по паре",
    "Контрагент", "Содержание", "Тип документа", "Риск (0-100)", "Причина"
]

    report_data["Дата"] = pd.to_datetime(report_data["Дата"]).dt.strftime("%d.%m.%Y %H:%M")
    report_data["Риск (0-100)"] = report_data["Риск (0-100)"].round(1)
    report_data["Сумма"] = report_data["Сумма"].round(2).astype(str)
    report_data["Средняя сумма по паре"] = report_data["Средняя сумма по паре"].round(2).astype(str)

    # Пишем в BytesIO вместо файла на диск
    buffer = io.BytesIO()
    report_data.to_excel(buffer, index=False, sheet_name="Аномалии")

    # Форматирование
    buffer.seek(0)
    wb = openpyxl.load_workbook(buffer)
    ws = wb["Аномалии"]

    red_fill    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[3].number_format = '#,##0.00'
        row[4].number_format = '#,##0.00'
        risk = row[8].value
        if risk and risk > 80:
            for cell in row:
                cell.fill = red_fill
        elif risk and risk > 60:
            for cell in row:
                cell.fill = yellow_fill

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    # Сохраняем обратно в BytesIO и возвращаем байты
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()